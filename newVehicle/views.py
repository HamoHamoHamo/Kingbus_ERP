from django.db.models import Sum  # Sum 함수 임포트
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from .models import Vehicle, Maintenance
from django.views.generic import ListView
import openpyxl
import json
from humanresource.models import Member
from .forms import VehicleForm, MaintenanceForm
from django.http import JsonResponse, HttpResponseBadRequest
from django.http import HttpResponseRedirect, HttpResponse
# Create your views here.
# def vehicle_view(request):
#     vehicles = Vehicle.objects.all()  # 데이터베이스에서 모든 차량 정보를 가져옵니다.
        
#     for vehicle in vehicles:
#         vehicle.details_count = vehicle.count_filled_fields()  # 값이 있는 필드 개수 계산
        


#     return render(request, 'newVehicle/vehicle_list.html', {'vehicles': vehicles})

class VehicleListView(ListView):
    model = Vehicle
    template_name = 'newVehicle/vehicle_list.html'
    context_object_name = 'vehicles'
    paginate_by = 10

    def get_queryset(self):
        # 데이터베이스에서 모든 차량 정보를 가져옵니다.
        vehicles = Vehicle.objects.all()

        
        
        # 각 차량 객체에 details_count 속성을 추가합니다.
        for vehicle in vehicles:
            vehicle.details_count = vehicle.count_filled_fields()

        return vehicles




    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        paginator = context['paginator']
        page_numbers_range = 5
        max_index = len(paginator.page_range)
        page = self.request.GET.get('page')
        current_page = int(page) if page else 1

        # 페이지 범위를 계산합니다.
        start_index = int((current_page - 1) / page_numbers_range) * page_numbers_range
        end_index = start_index + page_numbers_range
        if end_index >= max_index:
            end_index = max_index
        page_range = paginator.page_range[start_index:end_index]
        context['page_range'] = page_range

        # 각 페이지의 시작 번호를 정확하게 계산하고 1을 더해줍니다.
        context['start_num'] = paginator.per_page * (current_page - 1) + 1

        # 운전원, 팀장, 용역인 멤버 가져오기 (driver_list 추가)
        driver_list = Member.objects.filter(role__in=['운전원', '팀장', '용역'])
        print(driver_list)  # 콘솔에 driver_list 출력하여 값 확인
        context['driver_list'] = driver_list
        
        maintenance_records = Maintenance.objects.all()
        context['maintenance_records'] = maintenance_records


        return context






def vehicle_create(request):
    
    if request.method == 'POST':
        form = VehicleForm(request.POST)
        if form.is_valid(): #유효성 검사
            form.save()
            # URL name을 기반으로 리다이렉트
            return HttpResponseRedirect(reverse('newVehicle:vehicle_view'))
        else:
            return HttpResponseRedirect(reverse('newVehicle:vehicle_view'))
    else:
        form = VehicleForm()
        
    
    return HttpResponseRedirect(reverse('newVehicle:vehicle_view'))

    
def vehicle_delete_multiple(request):
    if request.method == 'POST':
        vehicle_ids = request.POST.getlist('vehicle_ids')
        Vehicle.objects.filter(id__in=vehicle_ids).delete()
        return redirect('newVehicle:vehicle_view')
    
def vehicle_update(request, pk):
    vehicle = get_object_or_404(Vehicle, id=pk)

    if request.method == 'POST':
        form = VehicleForm(request.POST, request.FILES, instance=vehicle)
        if form.is_valid():
            form.save()
            # 성공적으로 저장된 경우, 차량 리스트로 리다이렉트
            return redirect('newVehicle:vehicle_view')
        else:
            return render(request, 'newVehicle/vehicle_list.html', {'form': form, 'vehicle': vehicle})
    else:
        form = VehicleForm(instance=vehicle)

    return render(request, 'newVehicle/vehicle_list.html', {'form': form, 'vehicle': vehicle})



def vehicle_redirect_to_list(request, pk):
    return redirect('newVehicle:vehicle_view')

def vehicle_download(request):
    # 엑셀 파일 생성
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Vehicle List"

    # 헤더 추가
    headers = ['차량번호 앞', '차량번호 뒤', '제조사', '차량 이름', '담당 기사', '차대번호', '승차 인원', '모델 연도']
    sheet.append(headers)

    # 차량 데이터 추가
    vehicles = Vehicle.objects.all()
    for vehicle in vehicles:
        sheet.append([vehicle.vehicle_number_front, vehicle.vehicle_number_back, vehicle.maker, vehicle.vehicle_name,
                      vehicle.driver, vehicle.vehicle_serial, vehicle.passenger_capacity, vehicle.model_year])

    # 엑셀 파일을 HttpResponse로 반환
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=vehicle_list.xlsx'
    workbook.save(response)
    return response

def vehicle_excel_upload(request):
    if request.method == 'POST' and request.FILES.get('excelUploadFile'):
        excel_file = request.FILES['excelUploadFile']
        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            # 엑셀 파일의 각 행을 순회하여 데이터베이스에 저장
            for row in sheet.iter_rows(min_row=2, values_only=True):
                Vehicle.objects.create(
                    vehicle_number_front=row[0],
                    vehicle_number_back=row[1],
                    maker=row[2],
                    vehicle_name=row[3],
                    driver=row[4],
                    vehicle_serial=row[5],
                    passenger_capacity=row[6],
                    model_year=row[7]
                )

            # 성공 후 리스트 페이지로 리다이렉트
            return HttpResponseRedirect('/newVehicle/list/')
        except Exception as e:
            # 에러 발생 시 JSON 응답으로 에러 메시지 반환
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    # 파일이 없거나 POST 요청이 아닌 경우
    return JsonResponse({'success': False, 'message': '엑셀 파일 업로드에 실패했습니다.'}, status=400)


def some_view(request):
    response_data = {
        'success': True,
        'message': '파일이 성공적으로 업로드되었습니다.'
    }
    # 한글이 깨지지 않도록 ensure_ascii=False 사용
    return JsonResponse(json.loads(json.dumps(response_data, ensure_ascii=False)), safe=False)


def maintenance_create_or_delete(request, vehicle_id):
    vehicle = get_object_or_404(Vehicle, id=vehicle_id)

    if request.method == 'POST':
        if request.POST.get('action') == 'delete':  # 삭제 처리
            maintenance_ids = request.POST.getlist('maintenance_ids')
            for maintenance_id in maintenance_ids:
                maintenance = get_object_or_404(Maintenance, id=maintenance_id)
                maintenance.delete()

        elif request.POST.get('action') == 'create':  # 등록 처리
            form = MaintenanceForm(request.POST)
            if form.is_valid():
                maintenance = form.save(commit=False)
                maintenance.vehicle = vehicle
                maintenance.save()

    # 총 정비 및 튜닝 금액 업데이트
    total_maintenance_cost = Maintenance.objects.filter(vehicle=vehicle, type='정비').aggregate(Sum('cost'))['cost__sum'] or 0
    total_tuning_cost = Maintenance.objects.filter(vehicle=vehicle, type='튜닝').aggregate(Sum('cost'))['cost__sum'] or 0
    vehicle.total_maintenance_cost = total_maintenance_cost
    vehicle.total_tuning_cost = total_tuning_cost
    vehicle.save()

    # 리다이렉트 사용
    return redirect('newVehicle:vehicle_view')  # 'vehicle_view' URL로 리다이렉트